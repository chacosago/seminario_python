# -necesito saber que alumnos de la lista de aweb están no estan en la lista de siu
# listas ubicadas en el mismo directorio
# -2do quienes están en siu pero no en au

# La idea seria sacar una lista de diccionarios, cada dicc es un alumno del cual 
# guardo: legajo, apellido, nombre, mail.



# import openpyxl
import xlrd
import sys
import os
import json
import openpyxl

# descomentar en vscode
#os.chdir("/home/hcastorp/Drive/Facultad/Informatica/Python/TP/otras-pruebas/proyecto-carla/")
print("Tenga en cuenta que los archivos a utilizar deben estar en la misma carpeta que el programa y deben incluir su extensión.\n")
archivo_aweb = input("Ingresar nombre del arhivo descargado de AulasWeb: ")
archivo_siu = input("Ingresar nombre del arhivo descargado de Siu-Guarani: ")

#excel_siu = xlrd.open_workbook('Inscripciones a cursadas.xls')
excel_siu = xlrd.open_workbook(archivo_siu)
hoja_siu = excel_siu.sheet_by_index(0)

alumnos_siu = []

for i in range(11,hoja_siu.nrows):
    fila = hoja_siu.row(i)
    apellido_nombre = hoja_siu.cell_value(i,1).split(",")
    alumnos_siu.append({"Legajo": hoja_siu.cell_value(i,0), "Apellido": apellido_nombre[0].lower(),
        "Nombre": apellido_nombre[1].lower(), "Mail": hoja_siu.cell_value(i,4)})


#########################################
# Ahora con los de AulasWeb

#excel_auweb = openpyxl.load_workbook('TV2_COM2_2021  ELECCION DE GRUPOS.xlsx')
excel_auweb = openpyxl.load_workbook(archivo_aweb)
hoja_auweb = excel_auweb.active

alumnos_auweb = []

for fila in hoja_auweb.iter_rows(values_only=True):
    alumnos_auweb.append({"Apellido": fila[0].lower(), "Nombre": fila[1].lower(), "Mail": fila[3], "Comision" : fila[4]})

####################################
# Resuelvo problema: qué alumnos de la lista de aweb están no estan en la lista de siu.

alumnos_no_encontrados = []

for alumno_w in alumnos_auweb:
    # me fijo si esta en siu:
    for alumno_s in alumnos_siu:
        esta = alumno_w["Mail"] == alumno_s["Mail"] or ( alumno_w["Nombre"] in alumno_s["Nombre"] and alumno_w["Apellido"] in alumno_s["Apellido"])
        if esta:
           break
    if not esta:
        alumnos_no_encontrados.append(alumno_w)


alumnos_no_encontrados.pop(0)
lista_ordenada = list(sorted(alumnos_no_encontrados,key=lambda t : t["Apellido"]))
c = 0
for i in lista_ordenada:
    c+=1
    print(f'{c:3d} {i["Nombre"].capitalize():<30}{i.get("Apellido").capitalize():<30}{i.get("Mail"):<30}')

